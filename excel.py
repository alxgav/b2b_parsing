
import pandas as pd
import os, json, csv
import setting as s
from pprint import pprint
from openpyxl import load_workbook
from loguru import logger
from openpyxl import Workbook
from openpyxl.styles import *
from openpyxl.utils import get_column_letter
import mailto as email
from datetime import  timedelta, datetime



path = os.path.dirname(os.path.realpath(__file__))
logger.add(f'{path}/excel_error.log', format= '{time} {level} {message}', level='DEBUG', serialize=False)
config_data = json.load(open(f'{path}/config.json', 'r'))

wb = load_workbook(''.join(s.getExcelFileName()))
wb.save('out/temp.xlsx')
data_b2b_excel = pd.read_excel('out/temp.xlsx', usecols=['SKU', 
                                                            'QTA RIMANENTE',
                                                            'BRUTTOMENGE', 
                                                            'BESTÄTIGTER LIEFERTERMIN',
                                                            'BESCHREIBUNG MODELL',
                                                            'BESCHREIBUNG FARBE',
                                                            'BESCHREIBUNG GRÖSSE',
                                                            'PREIS'],  engine='openpyxl', )
data_b2b_excel.to_csv('out/temp.csv', columns=['SKU',
                                                'QTA RIMANENTE',
                                                'BRUTTOMENGE', 
                                                'BESTÄTIGTER LIEFERTERMIN', 
                                                'BESCHREIBUNG MODELL', 
                                                'BESCHREIBUNG FARBE', 
                                                'BESCHREIBUNG GRÖSSE', 
                                                'PREIS'], index=False, sep=';')

detail_csv = s.read_csv(f'{path}/detail_bike.csv')
data_b2b = s.read_csv('out/temp.csv')
active_csv = s.read_csv('in/Active_Orders.csv', encoding='ISO-8859-1') 
stock_csv = s.read_csv('in/Stock.csv', encoding='ISO-8859-1')
name_csv = s.read_csv('in/Name.csv', encoding='ISO-8859-1')




@logger.catch
def detail_bike(data):
    bike_data = []
    for bikes in detail_csv:
        for excel in data:
            if bikes[0] == excel['SKU']:
                bike_data.append({'sku': bikes[0],
                                'product': bikes[1],
                                'color': bikes[2],
                                'size': bikes[3],
                                'price': bikes[4],
                                'availabel': bikes[5],
                                'order': excel['ORDER'],
                                'fgz1': excel['FGZ1'],
                                'remaining_qty': excel['REMAINING QTY'],
                                'delivery_date': excel['DELIVERY DATE'],
                                'net_qty': excel['NET QTY'],
                                'active_orders': excel['ACTIVE_ORDERS'],
                                'SERVIZIO CORSA RADSPORT GMBH':'SERVIZIO CORSA RADSPORT GMBH',
                                'I99':'I99',
                                '50107393':'50107393'
                                })
    return bike_data          

def create_all(bike_data):
    REMAINING_QTY = 0
    ACTIVE_ORDERS = 0
    FGZ1 = 0
    ORDER = 0
    DELIVERY_DATE = ''
    bike_id = [i['sku'] for i in bike_data]
    names_data = []
    for n in name_csv:
        if n[0] not in  bike_id:
            for active in active_csv:
                if n[0] == active[0]:
                    ACTIVE_ORDERS = int(float(active[1].replace(',', '.')))
            for fgz in stock_csv:
                if n[0] == fgz[0]:
                    FGZ1 = int(float(fgz[1].replace(',', '.')))
        for b2b in data_b2b:
            if n[0] == b2b[0]:
                try:
                    REMAINING_QTY = b2b[1]
                except:
                    REMAINING_QTY = 0
                try:
                    ORDER = b2b[2]
                except:
                    ORDER = 0
                try:
                    DELIVERY_DATE = b2b[3]
                except:
                    DELIVERY_DATE = ''

        names_data.append({'sku': n[0].strip(),
                    'product': n[1].strip(),
                    'color': n[2].strip(),
                    'size': n[3].strip(),
                    'price': n[4],
                    'availabel': "NO",
                    'order': ORDER,
                    'fgz1': FGZ1,
                    'remaining_qty': REMAINING_QTY,
                    'delivery_date': DELIVERY_DATE,
                    'net_qty': FGZ1 + int(REMAINING_QTY) -ACTIVE_ORDERS,
                    'active_orders': ACTIVE_ORDERS,
                    'SERVIZIO CORSA RADSPORT GMBH':'SERVIZIO CORSA RADSPORT GMBH',
                    'I99':'I99',
                    '50107393':'50107393'})
               
    print(len(names_data))
    join_list = [*bike_data, *names_data]
    print (len(join_list))
    df = pd.DataFrame(join_list)
    df.sort_values(by="sku", ascending = True, inplace =True)
    # df.drop_duplicates(inplace=True)
    df.drop_duplicates(subset='sku', inplace=True)
   
    df.to_csv('data.csv', index=False, sep = ';')
    # # return data
def create_data2():
    data_csv = s.read_csv(f'{path}/data.csv')
    new_data = []
    d_t = datetime.today()
    
    for n in data_csv:
        if n[9] != '':
            d_n = datetime.strptime(n[9], '%Y-%m-%d')
            # if d_n >= d_t:
            #     print (d_t, d_n) 
            # print (d_t, d_n)
        if int(n[8]) > 0 and int(n[10]) > 0 and d_n >= d_t:
            new_data.append({'sku': n[0],
                    'product': n[1],
                    'color': n[2],
                    'size': n[3],
                    'price': n[4],
                    'availabel':n[5],
                    'order': n[6],
                    'fgz1': n[7],
                    'remaining_qty': n[8],
                    'delivery_date': n[9],
                    'net_qty': n[10],
                    'active_orders': n[11],
                    'SERVIZIO CORSA RADSPORT GMBH':'SERVIZIO CORSA RADSPORT GMBH',
                    'I99':'I99',
                    '50107393':'50107393'})
    df = pd.DataFrame(new_data)
    df.to_csv('data1.csv', sep=';', index=False)


@logger.catch # merge b2b active_orders and stock
def merge_data():
    data_b2b = s.read_csv('out/temp.csv')
    data_excel = [] 
    ACTIVE_ORDERS = 0 
    FGZ1 = 0
    for data in data_b2b:
        SKU = data[0]
        REMAINING_QTY = data[1]
        ORDER = data[2]
        DELIVERY_DATE = data[3]
        PRODUCT = data[4]
        COLOR = data[5]
        SIZE = data[6]
        PRICE = data[7]
        
        if SKU[0:4] not in config_data['out_of_list'] and SKU[0:5] not in config_data['out_of_list']:
            for active in active_csv:
                if SKU in active[0]:
                    ACTIVE_ORDERS = int(float(active[1].replace(',', '.')))
            for fgz in stock_csv:
                if SKU in fgz[0]:
                    FGZ1 = int(float(fgz[1].replace(',', '.')))
            NET_QTY = FGZ1 + int(REMAINING_QTY) - ACTIVE_ORDERS 
            
            data_excel.append({'SKU':SKU, 
                        'REMAINING QTY': REMAINING_QTY,
                        'ORDER': ORDER,
                        'DELIVERY DATE': DELIVERY_DATE,
                        'ACTIVE_ORDERS': ACTIVE_ORDERS,
                        'FGZ1': FGZ1,
                        'NET QTY': NET_QTY,
                        'PRODUCT': PRODUCT,
                        'COLOR': COLOR,
                        'SIZE': SIZE,
                        'PRICE': PRICE})
    return data_excel
    
    # df = pd.DataFrame(data_excel)
    # df.sort_values(by="SKU", ascending = True, inplace =True)
    # df.to_csv('excel.csv', index=False, sep = ';')

#  excel part

def create_excel_yes():
    data_csv = s.read_csv(f'{path}/data.csv')
    
    wb = Workbook()

    header = ['sku',
              'product',
              'color',
              'size',
              'price',
              'availabel',
            #   'order',
              'ACTIVE_ORDERS',
              'FGZ1',
              'REMAINING QTY',
              'NET QTY',
              'DELIVERY DATE'
              ]

    thin_border = Border(left=Side(style='thin'), 
                            right=Side(style='thin'), 
                            top=Side(style='thin'), 
                            bottom=Side(style='thin'))
    target = wb.active
    
    row = 2
    col = 1
    # HEADER
    target.row_dimensions[1].height = 40
    for i in header:
        target.cell(1, col).value = i
        target.cell(1, col).border = thin_border
        target.cell(1,10).fill = PatternFill(start_color="A9331A", end_color="A9331A", fill_type='solid')
        target.cell(1,11).fill = PatternFill(start_color="A9331A", end_color="A9331A", fill_type='solid')
        
        col += 1
    # end HEADER
    for i in range(1,12):
        target.column_dimensions[get_column_letter(i)].auto_size = True
        target.cell(1, i).alignment = Alignment(wrap_text=True,horizontal='center', vertical='center')
    target.column_dimensions['B'].width = 40

    # f = open('data.csv', 'r')
    # csv_reader = csv.DictReader(f)
    
    for r in data_csv:
        # print (r[5])
        if r[5] == 'YES':
            target.cell(row, 1).value = r[0]
            target.cell(row, 1).alignment = Alignment(horizontal='center', vertical='center')
            target.row_dimensions[row].height = 40
            target.cell(row, 2).value = r[1]
            target.cell(row, 2).alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
            target.cell(row, 3).value = r[2]
            target.cell(row, 3).alignment = Alignment(wrap_text=True)
            target.cell(row, 4).value = r[3]
            target.cell(row, 4).alignment = Alignment(horizontal='center', vertical='center')
            target.cell(row, 5).value = r[4]
            target.cell(row, 5).alignment = Alignment(horizontal='center', vertical='center')
            
            target.cell(row, 6).value = r[5]

            target.cell(row, 6).alignment = Alignment(horizontal='center', vertical='center')
            target.cell(row, 7).value = r[11]

            target.cell(row, 7).alignment = Alignment(horizontal='center', vertical='center')
            target.cell(row, 8).value = r[7]
            target.cell(row, 8).alignment = Alignment(horizontal='center', vertical='center')
            target.cell(row, 9).value = int(float(r[8]))
            target.cell(row, 9).alignment = Alignment(horizontal='center', vertical='center')
            net_qty = int(float(r[10]))
            # try:
            #     net_qty = r['fgz1'] + r['remaining_qty'] - r['active_orders']
            # except:
            #     net_qty = 0
            target.cell(row, 10).value = net_qty
            if net_qty < 0:
                target.cell(row, 10).fill = PatternFill(start_color="FF0040", end_color="FF0040", fill_type='solid')
            target.cell(row, 10).alignment = Alignment(horizontal='center', vertical='center')
            if int(float(r[8])) == 0:
                target.cell(row, 11).value = ''
            else:
                target.cell(row, 11).value = r[9]
                target.cell(row, 11).alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
            
            row +=1
    wb.save('data_yes.xlsx')

def create_excel():
    data_csv = s.read_csv(f'{path}/data.csv')
    wb = Workbook()

    header = ['sku',
              'product',
              'color',
              'size',
              'price',
              'availabel',
              'ACTIVE_ORDERS',
              'FGZ1',
              'REMAINING QTY',
              'NET QTY',
              'DELIVERY DATE'
              ]

    thin_border = Border(left=Side(style='thin'), 
                            right=Side(style='thin'), 
                            top=Side(style='thin'), 
                            bottom=Side(style='thin'))
    target = wb.active
    
    row = 2
    col = 1
    # HEADER
    target.row_dimensions[1].height = 40
    for i in header:
        target.cell(1, col).value = i
        target.cell(1, col).border = thin_border
        target.cell(1,10).fill = PatternFill(start_color="A9331A", end_color="A9331A", fill_type='solid')
        target.cell(1,11).fill = PatternFill(start_color="A9331A", end_color="A9331A", fill_type='solid')
        
        col += 1
    # end HEADER
    for i in range(1,12):
        target.column_dimensions[get_column_letter(i)].auto_size = True
        target.cell(1, i).alignment = Alignment(wrap_text=True,horizontal='center', vertical='center')
    target.column_dimensions['B'].width = 40

    
    for r in data_csv:
        target.cell(row, 1).value = r[0]
        target.cell(row, 1).alignment = Alignment(horizontal='center', vertical='center')
        target.row_dimensions[row].height = 40
        target.cell(row, 2).value = r[1]
        target.cell(row, 2).alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
        target.cell(row, 3).value = r[2]
        target.cell(row, 3).alignment = Alignment(wrap_text=True)
        target.cell(row, 4).value = r[3]
        target.cell(row, 4).alignment = Alignment(horizontal='center', vertical='center')
        target.cell(row, 5).value = r[4]
        target.cell(row, 5).alignment = Alignment(horizontal='center', vertical='center')
        
        target.cell(row, 6).value = r[5]

        target.cell(row, 6).alignment = Alignment(horizontal='center', vertical='center')
        # active = r[11]
        # if active == '':
        #     active = 0
        target.cell(row, 7).value = r[11]

        target.cell(row, 7).alignment = Alignment(horizontal='center', vertical='center')
        target.cell(row, 8).value = r[7]
        target.cell(row, 8).alignment = Alignment(horizontal='center', vertical='center')
        try:
            target.cell(row, 9).value = int(float(r[8]))
        except:
            target.cell(row, 9).value = 0
        target.cell(row, 9).alignment = Alignment(horizontal='center', vertical='center')
        try:
            # net_qty = int(float(r['fgz1'])) + int(float(r['remaining_qty'])) - int(float(r['active_orders']))
            net_qty = int(float(r[10]))
            if net_qty < 0:
                target.cell(row, 10).fill = PatternFill(start_color="FF0040", end_color="FF0040", fill_type='solid')
        except:
            net_qty = 0
        target.cell(row, 10).value = net_qty
        
        target.cell(row, 10).alignment = Alignment(horizontal='center', vertical='center')
        # try:
        #     view_data = int(float(r['remaining_qty']))
        # except:
        #     view_data = 0
        # if view_data == 0:
        #     target.cell(row, 11).value = ''
        # else:
        target.cell(row, 11).value = r[9]
        target.cell(row, 11).alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
        
        row +=1
    wb.save('data_all.xlsx')

if __name__ in "__main__":
    
    SUBJECT = 'Bianchi Rennrad Stock Data File'
    FROM = 'alxgav@gmail.com'
    html = 'file'
    create_all(detail_bike(merge_data()))
    create_excel_yes()
    create_excel()
    create_data2()
    print ('prepare email')

    email.send_mail_attachment(SUBJECT, FROM, config_data['email'], html)

    

    